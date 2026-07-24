# =================================================
# MODELS USERS — VERSION CORRIGÉE
# Matricule uniquement dans OfficialAuthority
# Toutes les infos personnelles saisies par l'autorité lors de l'inscription
# =================================================

from django.contrib.auth.models import AbstractUser
from django.db import models
import uuid
import os
import logging

logger = logging.getLogger(__name__)


class User(AbstractUser):
    """Custom User model with roles for Care-Link RDC."""

    ROLE_CHOICES = [
        ("citizen", "Citoyen"),
        ("hospital", "Hôpital"),
        ("authority", "Autorité"),
        ("admin", "Administrateur"),
    ]

    GENDER_CHOICES = [
        ("M", "Masculin"),
        ("F", "Féminin"),
    ]

    MARITAL_STATUS_CHOICES = [
        ("Célibataire", "Célibataire"),
        ("Marié(e)", "Marié(e)"),
        ("Divorcé(e)", "Divorcé(e)"),
        ("Veuf/Veuve", "Veuf/Veuve"),
    ]

    id = models.UUIDField(primary_key=True, default=uuid.uuid4, editable=False)
    email = models.EmailField(unique=True)
    role = models.CharField(max_length=20, choices=ROLE_CHOICES, default="citizen")

    # Profile fields
    phone = models.CharField(max_length=20, blank=True)
    avatar = models.ImageField(upload_to="avatars/", blank=True, null=True)

    # Personal info (for citizens)
    date_of_birth = models.DateField(blank=True, null=True)
    gender = models.CharField(max_length=1, choices=GENDER_CHOICES, blank=True)
    marital_status = models.CharField(
        max_length=20, choices=MARITAL_STATUS_CHOICES, blank=True
    )
    nationality = models.CharField(max_length=50, default="Congolaise")

    # Address
    province = models.CharField(max_length=100, blank=True)
    commune = models.CharField(max_length=100, blank=True)
    city = models.CharField(max_length=100, blank=True, verbose_name="Ville")
    address = models.TextField(blank=True)

    # Additional citizen info
    birth_place = models.CharField(
        max_length=255, blank=True, verbose_name="Lieu de naissance"
    )
    id_number = models.CharField(
        max_length=50, blank=True, verbose_name="Numéro d'identité"
    )

    # Professional info (for hospital/authority)
    # NOTE: pour les autorités, matricule = clé de vérification via OfficialAuthority
    matricule = models.CharField(max_length=50, blank=True, unique=True, null=True)
    grade = models.CharField(max_length=100, blank=True)
    service = models.CharField(max_length=200, blank=True)
    direction = models.CharField(max_length=200, blank=True)

    # Hospital relation (for hospital role)
    hospital = models.ForeignKey(
        "hospitals.Hospital",
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name="staff",
    )

    # Settings
    language = models.CharField(max_length=5, default="fr")
    theme = models.CharField(max_length=10, default="system")
    notifications_enabled = models.BooleanField(default=True)

    # Password
    password_reset_code = models.CharField(max_length=6, blank=True, null=True)
    password_reset_code_expires = models.DateTimeField(blank=True, null=True)

    # Timestamps
    created_at = models.DateTimeField(auto_now_add=True)
    updated_at = models.DateTimeField(auto_now=True)
    last_login_ip = models.GenericIPAddressField(blank=True, null=True)

    USERNAME_FIELD = "email"
    REQUIRED_FIELDS = ["username", "first_name", "last_name"]

    class Meta:
        db_table = "users"
        verbose_name = "Utilisateur"
        verbose_name_plural = "Utilisateurs"
        ordering = ["-created_at"]

    def __str__(self):
        return f"{self.first_name} {self.last_name} ({self.email})"

    @property
    def full_name(self):
        return f"{self.first_name} {self.last_name}".strip()

    @property
    def is_citizen(self):
        return self.role == "citizen"

    @property
    def is_hospital(self):
        return self.role == "hospital"

    @property
    def is_authority(self):
        return self.role == "authority"

    @property
    def is_admin(self):
        return self.role == "admin"


# =================================================
# OfficialAuthority — SEULEMENT le matricule
# Le reste des infos (nom, grade, service, email...)
# est saisi par l'autorité elle-même lors de l'inscription (dans User)
# =================================================


class OfficialAuthority(models.Model):
    """Liste officielle des matricules d'autorités autorisées.

    Contient UNIQUEMENT le matricule. Aucune info personnelle.
    L'autorité saisit elle-même ses données lors de l'inscription (User).
    """

    id = models.UUIDField(primary_key=True, default=uuid.uuid4, editable=False)
    matricule = models.CharField(max_length=50, unique=True)
    is_active = models.BooleanField(default=True)
    used_for_registration = models.BooleanField(default=False)
    created_at = models.DateTimeField(auto_now_add=True)

    class Meta:
        db_table = "official_authorities"
        ordering = ["matricule"]

    def __str__(self):
        return self.matricule

    @classmethod
    def verify_matricule(cls, matricule):
        """Vérifie si un matricule existe et est disponible pour inscription.

        Returns:
            dict: {
                'valid': bool,
                'reason': str (si invalid),
                'matricule': str (si valid)
            }
        """
        if not matricule:
            return {"valid": False, "reason": "Le matricule est requis."}

        matricule = matricule.strip().upper()
        logger.info(f"Vérification matricule: '{matricule}'")

        # ÉTAPE 1: Chercher dans la base de données
        try:
            authority = cls.objects.get(matricule=matricule)
            logger.info(f"Matricule trouvé en DB: {matricule}")

            if not authority.is_active:
                return {
                    "valid": False,
                    "reason": "Ce matricule est désactivé. Contactez l'administrateur système.",
                }

            if authority.used_for_registration:
                return {
                    "valid": False,
                    "reason": "Ce matricule a déjà été utilisé pour créer un compte.",
                }

            return {"valid": True, "matricule": authority.matricule}

        except cls.DoesNotExist:
            logger.info(
                f"Matricule non trouvé en DB: {matricule}. Recherche dans Excel..."
            )

        # ÉTAPE 2: Chercher dans le fichier Excel/CSV (fallback)
        excel_path = cls.get_excel_path()
        if excel_path:
            try:
                found = cls._search_in_excel(matricule, excel_path)
                if found:
                    logger.info(f"Matricule trouvé dans Excel: {matricule}")

                    # Ajouter automatiquement en DB pour les prochaines fois
                    try:
                        cls.objects.create(matricule=matricule)
                        logger.info(
                            f"Matricule ajouté automatiquement en DB: {matricule}"
                        )
                    except Exception as e:
                        logger.warning(f"Impossible d'ajouter en DB: {e}")

                    return {"valid": True, "matricule": matricule}

            except Exception as e:
                logger.error(f"Erreur lors de la recherche Excel: {e}")

        logger.warning(f"Matricule introuvable (DB + Excel): {matricule}")
        return {
            "valid": False,
            "reason": "Matricule non reconnu. Vérifiez votre numéro ou contactez l'administrateur système.",
        }

    # =================================================
    # CHEMIN DU FICHIER EXCEL/CSV
    # =================================================

    @classmethod
    def get_excel_path(cls):
        """Retourne le chemin absolu du fichier Excel/CSV.
        Cherche dans plusieurs emplacements possibles.
        """
        possible_paths = [
            os.path.join(
                os.path.dirname(__file__), "data", "official_authorities.xlsx"
            ),
            os.path.join(os.path.dirname(__file__), "data", "official_authorities.csv"),
            os.path.join(
                os.path.dirname(__file__), "fixtures", "official_authorities.xlsx"
            ),
            os.path.join(
                os.path.dirname(__file__), "fixtures", "official_authorities.csv"
            ),
            os.path.join(
                os.path.dirname(__file__), "..", "data", "official_authorities.xlsx"
            ),
            os.path.join(
                os.path.dirname(__file__), "..", "data", "official_authorities.csv"
            ),
            os.path.join(
                os.path.dirname(__file__), "..", "..", "official_authorities.xlsx"
            ),
            os.path.join(
                os.path.dirname(__file__), "..", "..", "official_authorities.csv"
            ),
        ]

        env_path = os.environ.get("OFFICIAL_AUTHORITIES_EXCEL_PATH")
        if env_path:
            possible_paths.insert(0, env_path)

        for path in possible_paths:
            abs_path = os.path.abspath(path)
            if os.path.exists(abs_path):
                logger.info(f"Fichier Excel trouvé: {abs_path}")
                return abs_path

        logger.error("Aucun fichier Excel trouvé dans les chemins connus!")
        return None

    # =================================================
    # RECHERCHE DANS EXCEL/CSV (SIMPLIFIÉ — matricule uniquement)
    # =================================================

    @classmethod
    def _search_in_excel(cls, matricule, filepath):
        """Cherche un matricule dans le fichier Excel/CSV.
        Le fichier doit contenir une colonne 'matricule'.
        Retourne True si trouvé, False sinon.
        """
        matricule = matricule.strip().upper()

        if filepath.endswith(".csv"):
            return cls._search_in_csv(matricule, filepath)

        # === EXCEL (.xlsx) ===
        try:
            import openpyxl
        except ImportError:
            logger.error(
                "openpyxl n'est pas installé. Installez-le: pip install openpyxl"
            )
            return False

        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
        except Exception as e:
            logger.error(f"Erreur ouverture fichier Excel {filepath}: {e}")
            return False

        # Détection de la ligne d'en-tête (colonne 'matricule')
        header_row_idx = None
        matricule_col_idx = None
        max_rows_to_check = 5

        for row_idx in range(1, min(max_rows_to_check + 1, ws.max_row + 1)):
            for col_idx, cell in enumerate(ws[row_idx], start=1):
                if cell.value and str(cell.value).strip().lower() == "matricule":
                    header_row_idx = row_idx
                    matricule_col_idx = col_idx
                    logger.info(
                        f"En-tête trouvé à la ligne {row_idx}, colonne {col_idx}"
                    )
                    break
            if header_row_idx:
                break

        if header_row_idx is None:
            logger.error(
                f"Colonne 'matricule' non trouvée dans les {max_rows_to_check} premières lignes."
            )
            return False

        # Parcourir les lignes de données
        for row in ws.iter_rows(
            min_row=header_row_idx + 1,
            min_col=matricule_col_idx,
            max_col=matricule_col_idx,
            values_only=True,
        ):
            if row and row[0]:
                row_matricule = str(row[0]).strip().upper()
                if row_matricule == matricule:
                    logger.info(f"Matricule {matricule} trouvé dans Excel")
                    return True

        logger.info(f"Matricule {matricule} non trouvé dans Excel")
        return False

    @classmethod
    def _search_in_csv(cls, matricule, filepath):
        """Cherche un matricule dans un fichier CSV."""
        import csv

        try:
            with open(filepath, "r", encoding="utf-8") as f:
                reader = csv.reader(f)

                # Détection de l'en-tête
                header_row_idx = None
                matricule_col_idx = None

                for row_idx, row in enumerate(reader, start=1):
                    if row_idx > 5:
                        break

                    for col_idx, cell in enumerate(row, start=1):
                        if cell and str(cell).strip().lower() == "matricule":
                            header_row_idx = row_idx
                            matricule_col_idx = col_idx
                            logger.info(f"En-tête CSV trouvé à la ligne {row_idx}")
                            break
                    if header_row_idx:
                        break

                if header_row_idx is None:
                    logger.error("En-tête CSV non trouvé")
                    return False

                # Re-ouvrir et lire les données
                f.seek(0)
                reader = csv.reader(f)

                # Skip jusqu'à l'en-tête
                for _ in range(header_row_idx):
                    next(reader, None)

                # Lire les données
                for row in reader:
                    if len(row) >= matricule_col_idx:
                        row_matricule = (
                            str(row[matricule_col_idx - 1] or "").strip().upper()
                        )
                        if row_matricule == matricule:
                            logger.info(f"Matricule {matricule} trouvé dans CSV")
                            return True

        except Exception as e:
            logger.error(f"Erreur lecture CSV {filepath}: {e}")

        return False

    # =================================================
    # IMPORT EXCEL/CSV (SIMPLIFIÉ — matricule uniquement)
    # =================================================

    @classmethod
    def load_from_excel(cls, filepath=None):
        """Charger les matricules depuis un fichier Excel (.xlsx).
        Le fichier doit contenir une colonne 'matricule'.
        """
        if filepath is None:
            filepath = cls.get_excel_path()

        if not filepath:
            raise FileNotFoundError("Aucun fichier Excel trouvé.")

        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl est requis. Installez-le avec: pip install openpyxl"
            )

        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Fichier non trouvé: {filepath}")

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Détection de l'en-tête
        header_row_idx = None
        matricule_col_idx = None

        for row_idx in range(1, min(6, ws.max_row + 1)):
            for col_idx, cell in enumerate(ws[row_idx], start=1):
                if cell.value and str(cell.value).strip().lower() == "matricule":
                    header_row_idx = row_idx
                    matricule_col_idx = col_idx
                    break
            if header_row_idx:
                break

        if header_row_idx is None:
            raise ValueError(
                "Colonne 'matricule' non trouvée dans les 5 premières lignes."
            )

        created_count = 0
        errors = []

        for row in ws.iter_rows(
            min_row=header_row_idx + 1,
            min_col=matricule_col_idx,
            max_col=matricule_col_idx,
            values_only=True,
        ):
            if row and row[0]:
                matricule = str(row[0]).strip().upper()
                if not matricule:
                    continue

                try:
                    obj, created = cls.objects.get_or_create(matricule=matricule)
                    if created:
                        created_count += 1
                except Exception as e:
                    errors.append(f"{matricule}: {str(e)}")

        return {"created": created_count, "errors": errors}

    @classmethod
    def load_from_csv(cls, filepath=None):
        """Charger les matricules depuis un fichier CSV.
        Le fichier doit contenir une colonne 'matricule'.
        """
        if filepath is None:
            filepath = cls.get_excel_path()
            if filepath and not filepath.endswith(".csv"):
                csv_path = filepath.replace(".xlsx", ".csv").replace(".xls", ".csv")
                if os.path.exists(csv_path):
                    filepath = csv_path
                else:
                    raise FileNotFoundError("Aucun fichier CSV trouvé.")

        if not filepath or not os.path.exists(filepath):
            raise FileNotFoundError(f"Fichier non trouvé: {filepath}")

        import csv

        created_count = 0
        errors = []

        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.reader(f)

            # Détection de l'en-tête
            header_row_idx = None
            matricule_col_idx = None

            for row_idx, row in enumerate(reader, start=1):
                if row_idx > 5:
                    break
                for col_idx, cell in enumerate(row, start=1):
                    if cell and str(cell).strip().lower() == "matricule":
                        header_row_idx = row_idx
                        matricule_col_idx = col_idx
                        break
                if header_row_idx:
                    break

            if header_row_idx is None:
                raise ValueError("En-tête non trouvé dans le CSV.")

            # Re-ouvrir pour lire les données
            f.seek(0)
            reader = csv.reader(f)

            # Skip jusqu'à l'en-tête
            for _ in range(header_row_idx):
                next(reader, None)

            for row in reader:
                if len(row) >= matricule_col_idx:
                    matricule = str(row[matricule_col_idx - 1] or "").strip().upper()
                    if not matricule:
                        continue

                    try:
                        obj, created = cls.objects.get_or_create(matricule=matricule)
                        if created:
                            created_count += 1
                    except Exception as e:
                        errors.append(f"{matricule}: {str(e)}")

        return {"created": created_count, "errors": errors}


class AuditLog(models.Model):
    """Audit trail for all system actions."""

    ACTION_CHOICES = [
        ("login", "Connexion"),
        ("login_failed", "Échec connexion"),
        ("logout", "Déconnexion"),
        ("create", "Création"),
        ("update", "Modification"),
        ("delete", "Suppression"),
        ("validate", "Validation"),
        ("reject", "Rejet"),
        ("export", "Export"),
        ("view", "Consultation"),
        ("payment", "Paiement"),
    ]

    id = models.UUIDField(primary_key=True, default=uuid.uuid4, editable=False)
    user = models.ForeignKey(User, on_delete=models.SET_NULL, null=True, blank=True)
    action = models.CharField(max_length=20, choices=ACTION_CHOICES)
    target_type = models.CharField(max_length=50)
    target_id = models.CharField(max_length=100)
    details = models.JSONField(default=dict)
    ip_address = models.GenericIPAddressField(blank=True, null=True)
    user_agent = models.TextField(blank=True)
    timestamp = models.DateTimeField(auto_now_add=True)

    class Meta:
        db_table = "audit_logs"
        ordering = ["-timestamp"]

    def __str__(self):
        return f"{self.action} by {self.user} at {self.timestamp}"
